from openpyxl import *
from icecream import ic
from openpyxl.worksheet import cell_range


class ExcelOp(object):
    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file, data_only=True)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]

    # 获取表格的总行数和总列数
    def get_row_clo_num(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def get_cell_value(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    # 获取某列的所有值
    def get_col_value(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    # 设置某个单元格的值
    def set_cell_value(self, row, colunm, cellvalue):
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file)


def table_ziZhu():
    excel_op = ExcelOp(file=r"D:\MyData\ex_pengzb\Downloads\自主维护记录修改新.xlsx")
    exe_code = excel_op.get_col_value(2)
    # ic(exe_code)
    device_code = excel_op.get_col_value(4)
    # ic(device_code)
    begin_date = excel_op.get_col_value(5)
    # ic(begin_date)
    end_date_before = excel_op.get_col_value(6)
    # ic(end_date_before)
    end_date_after = excel_op.get_col_value(7)
    # ic(end_date_after)

    # print(type(str(end_date_after[12])))
    # ic(str(end_date_after[12]))
    # ic(len(exe_code))
    exe_code_str_1 = ''
    for i in range(3, 1000):
        exe_code_str_1 += "'" + exe_code[i] + "'"
        if i < 1000 - 1:
            exe_code_str_1 += ","

    exe_code_str_2 = ''
    for i in range(1000, len(exe_code)):
        exe_code_str_2 += "'" + exe_code[i] + "'"
        if i < len(exe_code) - 1:
            exe_code_str_2 += ","

    # ic(exe_code_str_1)
    # ic(exe_code_str_2)
    insertSql = "insert into geam.pm_plan_his \n" \
                "  select sysdate as chg_date,\n" \
                "         '彭镇波' as chg_user_name,\n" \
                "         '修改' as chg_type,\n" \
                "         '部分自主维护实施单，专业点检实施单和专业润滑实施单的实际完工时间有误，因此申请数据修正实际完工时间' as chg_desc,\n" \
                "         pp.* \n" \
                "    from geam.pm_plan pp \n" \
                "   where pp.exe_code in (" + exe_code_str_1 + ");"

    # print(insertSql)

    for i in range(3, 1000):
        updateSql = "update pm_plan pp set pp.end_date = to_date('" + str(
            end_date_after[i]) + "','yyyy-MM-dd hh24:mi:ss') \n " \
                                 " where pp.exe_code = '" + exe_code[i] + "'" \
                                                                          " and pp.device_code = '" + device_code[
                        i] + "'" \
                             " and pp.end_date = to_date('" + str(end_date_before[i]) + "','yyyy-MM-dd hh24:mi:ss') " \
                                                                                        " and pp.begin_date = to_date('" + str(
            begin_date[i]) + "','yyyy-MM-dd hh24:mi:ss');"
        # print(updateSql)

    insertSql = "insert into geam.pm_plan_his \n" \
                "  select sysdate as chg_date,\n" \
                "         '彭镇波' as chg_user_name,\n" \
                "         '修改' as chg_type,\n" \
                "         '部分自主维护实施单，专业点检实施单和专业润滑实施单的实际完工时间有误，因此申请数据修正实际完工时间' as chg_desc,\n" \
                "         pp.* \n" \
                "    from geam.pm_plan pp \n" \
                "   where pp.exe_code in (" + exe_code_str_2 + ");"

    # print(insertSql)

    for i in range(1000, len(exe_code)):
        updateSql = "update pm_plan pp set pp.end_date = to_date('" + str(
            end_date_after[i]) + "','yyyy-MM-dd hh24:mi:ss') \n " \
                                 " where pp.exe_code = '" + exe_code[i] + "'" \
                                                                          " and pp.device_code = '" + device_code[
                        i] + "'" \
                             " and pp.end_date = to_date('" + str(end_date_before[i]) + "','yyyy-MM-dd hh24:mi:ss') " \
                                                                                        " and pp.begin_date = to_date('" + str(
            begin_date[i]) + "','yyyy-MM-dd hh24:mi:ss');"
        print(updateSql)

if __name__ == '__main__':
    excel_op = ExcelOp(file=r"D:\MyData\ex_pengzb\Downloads\自主维护记录修改新.xlsx")
    exe_code = excel_op.get_col_value(2)
    # ic(exe_code)
    device_code = excel_op.get_col_value(4)
    # ic(device_code)
    begin_date = excel_op.get_col_value(5)
    # ic(begin_date)
    end_date_before = excel_op.get_col_value(6)
    # ic(end_date_before)
    end_date_after = excel_op.get_col_value(7)
    # ic(end_date_after)


