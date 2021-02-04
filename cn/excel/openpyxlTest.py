from openpyxl import *
from icecream import ic
from openpyxl.worksheet import cell_range

wb = load_workbook(r"D:\PycharmProjects\demoProj\cn\excel\2020年损益表-宝华.xlsx", data_only=True)
sheet_names = wb.sheetnames
ws = wb[sheet_names[0]]  # index为0为第一张表
ic(ws)
d = ws.cell(row=79, column=3).value
ic(d)
