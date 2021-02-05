import time

import os
from icecream import ic
from openpyxl import *


def get_file_from_fold(_path):
    files = os.listdir(_path)
    files_list = []
    for filename in files:
        if not (filename.__contains__('.xls') or filename.__contains__('.xlsx')) or (filename.__contains__('output')):
            continue
        files_list.append(_path + filename)
    return files_list


class Excel_reader:
    """excel读取类"""
    _a_list = []
    _head_list = []

    def __init__(self, _file_name):
        self.file = _file_name
        self.wb = load_workbook(self.file, data_only=True)
        sheets = self.wb.sheetnames
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]
        self.row_sum, self.col_sum = self.get_row_col_sum()

    def get_sheet(self):
        """
        获取表
        :return:
        """
        # 获取所有sheet的名字
        sheet_names = self.wb.sheetnames
        name = self.file
        str_list = name.split('\\')  # 截取生成表名
        for _sheet_name in sheet_names:
            print('文件名：' + str_list[len(str_list) - 1] + ':\n表名：', _sheet_name)
        # sheet1索引从0开始，得到sheet1表
        sheet = sheet_names[0]
        time.sleep(1)
        return sheet

    # 获取表格的总行数和总列数
    def get_row_col_sum(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def get_cell_value(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    # 获取某列的所有值
    def get_col_value(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    # 寻找最后一行
    def find_last_row(self):
        _row_sum, _ = self.get_row_col_sum()
        row_val = [None] * _row_sum
        while all(x is None for x in row_val):
            _row_sum -= 1
            row_val = self.get_row_value(_row_sum)

        return _row_sum


class Excel_writer:
    """
    写入excel类
    """

    def __init__(self, _file_name):
        self.file = _file_name
        self.wb = Workbook()
        sheets = self.wb.sheetnames
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]

    def set_cell_value(self, row, column, cell_value):
        """
        设置某个单元格的值
        """
        try:
            self.ws.cell(row=row, column=column).value = cell_value
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=column).value = "write failed"
            self.wb.save(self.file)


def banner_paint():
    print("""
     ______     __  __     ______     ______     __            ______   ______     ______     __        
    /\  ___\   /\_\_\_\   /\  ___\   /\  ___\   /\ \          /\__  _\ /\  __ \   /\  __ \   /\ \       
    \ \  __\   \/_/\_\/_  \ \ \____  \ \  __\   \ \ \____     \/_/\ \/ \ \ \/\ \  \ \ \/\ \  \ \ \____  
     \ \_____\   /\_\/\_\  \ \_____\  \ \_____\  \ \_____\       \ \_\  \ \_____\  \ \_____\  \ \_____\ 
      \/_____/   \/_/\/_/   \/_____/   \/_____/   \/_____/        \/_/   \/_____/   \/_____/   \/_____/                                                                                              
    """)


input_fold_name = ''


# 初始化程序
def init():
    banner_paint()
    global input_fold_name
    input_fold_name = input("请输入需要汇总文件夹的路径：")
    input_fold_name += '\\'
    print("开始提取...")
    time.sleep(2)


# 判断变量类型的函数
def typeof(variate):
    _type = None
    if isinstance(variate, int):
        _type = "int"
    elif isinstance(variate, str):
        _type = "str"
    elif isinstance(variate, float):
        _type = "float"
    elif isinstance(variate, list):
        _type = "list"
    elif isinstance(variate, tuple):
        _type = "tuple"
    elif isinstance(variate, dict):
        _type = "dict"
    elif isinstance(variate, set):
        _type = "set"
    return _type


if __name__ == '__main__':
    # ..\excel\
    # ..\testExcel\
    # input_fold_name = '..\\testExcel\\'
    # input_fold_name = '..\\excel\\'
    init()
    print("获取输入的文件夹的excel文件...")
    a_list = get_file_from_fold(_path=input_fold_name)  # 获取文件夹内的所有含有.xls结尾的文件
    if len(a_list) == 0:
        print("err: 并没有找到文件夹含有.xls结尾的文件。\n请添加后重新运行此程序!")
        os.abort()
    # ic(a_list)
    time.sleep(1)
    print("======导入excel完成！需要合并的文件有", len(a_list), "个")
    time.sleep(2)
    row_counter = 0
    print("======新建output文件！")
    ew = Excel_writer(f'{input_fold_name}\output.xlsx')  # 新建output文件
    for i in range(len(a_list)):
        er = Excel_reader(_file_name=a_list[i])
        row_sum, col_sum = er.get_row_col_sum()
        real_row_sum = er.find_last_row()  # 查找最后一行 （最后一行有可能是None）
        # ic(real_row_sum)
        print('查找最后一行为：', real_row_sum)
        # 设置获取最后n行
        set_last_row_num = 3
        # 写入表名
        sheet_name = str(er.get_sheet())
        row_counter += 1
        ew.set_cell_value(row_counter, 1, sheet_name)
        # 写入表头
        row_counter += 1
        res_data_3 = er.get_row_value(3)
        for j in range(len(res_data_3)):
            print(j, ":", res_data_3[j])
            ew.set_cell_value(row_counter, j + 1, res_data_3[j])
            # 写入内容
        for jj in range(set_last_row_num):
            last_row_n = er.get_row_value(real_row_sum - jj)
            row_counter += 1
            for j in range(len(last_row_n)):
                print(j, ":", last_row_n[j])
                # ic(type(last_row_n[j]))
                if last_row_n[j] is None:
                    continue
                # 小数转化百分比显示
                if typeof(last_row_n[j]) == typeof(1) or typeof(last_row_n[j]) == typeof(0.1) and (
                        1 >= last_row_n[j] > 0):
                    last_row_n[j] = format(last_row_n[j], '.2%')
                    # ic(last_row_n[j])
                ew.set_cell_value(row_counter, j + 1, last_row_n[j])
